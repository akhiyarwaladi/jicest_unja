import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Read the CSV file - try different encodings
try:
    df = pd.read_csv('tests/ua1.csv', delimiter=';', encoding='utf-8')
except UnicodeDecodeError:
    try:
        df = pd.read_csv('tests/ua1.csv', delimiter=';', encoding='latin-1')
    except:
        df = pd.read_csv('tests/ua1.csv', delimiter=';', encoding='cp1252')

# Create Excel file
output_file = 'tests/ua1_formatted.xlsx'
df.to_excel(output_file, index=False, sheet_name='Abstract Submissions')

# Load workbook for formatting
wb = load_workbook(output_file)
ws = wb.active

# Define styles
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2C5F2D', end_color='2C5F2D', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

cell_font = Font(name='Calibri', size=10)
cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

border_style = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Apply header formatting
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border_style

# Set column widths and apply formatting to data cells
column_widths = {
    'A': 60,  # title
    'B': 40,  # authors
    'C': 35,  # institutions
    'D': 80,  # abstract
    'E': 40,  # keywords
    'F': 25,  # presenter
    'G': 25,  # full_name1
    'H': 25,  # full_name2
    'I': 20,  # participant_type
    'J': 40,  # address
    'K': 18,  # phone
    'L': 15   # attendance
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Apply formatting to all data cells
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border_style

# Set row height for header
ws.row_dimensions[1].height = 30

# Set row heights for data rows
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = None  # Auto-adjust based on content

# Freeze the header row
ws.freeze_panes = 'A2'

# Apply alternating row colors for better readability
for row in range(2, ws.max_row + 1):
    if row % 2 == 0:
        fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        for cell in ws[row]:
            cell.fill = fill

# Save the formatted workbook
wb.save(output_file)

print(f"✓ Excel file created successfully: {output_file}")
print(f"✓ Total submissions: {len(df)}")
print(f"✓ Columns: {', '.join(df.columns)}")
